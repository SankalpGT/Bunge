import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import re

def float_to_hhmm(hrs_float: float) -> str:
    # extract whole hours
    hours = int(hrs_float)
    # convert fractional part to minutes
    minutes = int(round((hrs_float - hours) * 60))
    # handle the case where rounding pushes minutes to 60
    if minutes == 60:
        hours += 1
        minutes = 0
    return f"{hours:02d}:{minutes:02d}"

def generate_excel_from_extracted_data(metadata: dict, deductions: list[dict], net_laytime_used_hours: float):
    wb = Workbook()
    ws = wb.active
    ws.title = "LAY TIME CALCULATIONS"

    # --- Calculate Laytime Allowed ---
    time_allowed = ""
    try:
        # Robustly extract numerical part from Quantity and Discharge Rate
        quantity_raw = metadata.get("Quantity", "")
        # Changed key from DISRATE to Discharge Rate
        disrate_raw = metadata.get("DISRATE", "") 

        # Use regex to find numbers (integers or floats)
        quantity_match = re.search(r'(\d+\.?\d*)', str(quantity_raw))
        disrate_match = re.search(r'(\d+\.?\d*)', str(disrate_raw))

        quantity = float(quantity_match.group(1)) if quantity_match else 0.0
        disrate = float(disrate_match.group(1)) if disrate_match else 0.0

        if disrate != 0:
            time_allowed = f"{quantity / disrate:.4f}"
        else:
            time_allowed = "N/A (Discharge Rate is zero)"
    except Exception as e:
        time_allowed = f"Error calculating Laytime Allowed: {e}"

    laytime_allowed_value = deductions[0].get('deducted_to', 'N/A Discharge rate is 0')

    a_c = ""
    if metadata.get("A/C", ""):
        a_c = metadata.get("A/C", "")
    else:
        a_c = metadata.get("Charterer", "")

    # -- Top: Metadata Header Rows --
    header_rows = [
        ["Vessel Name :", metadata.get("Vessel Name", ""), "", "", "PORT :", metadata.get("Port", "")],
        ["A/C :", a_c, "", "", "QUANTITY :", metadata.get("Quantity", "")],
        ["TERMS :", metadata.get("TERMS", ""), "DISRATE :", metadata.get("DISRATE", ""), "NOR TENDERED :", metadata.get("NOR TENDERED", "")], 
        ["PRODUCT :", metadata.get("PRODUCT", ""), "", "", "VESSEL ARRIVED :", metadata.get("Vessel Arrival", "")],
        ["LTC  AT :", metadata.get("LTC AT", ""), "", "", "VESSEL BERTHED :", metadata.get("Vessel Berthed", "")],
        ["DEMMURAGE :", metadata.get("DEMMURAGE", ""), "", "", "COMMENCED CARGO :", metadata.get("Commenced Cargo", "")],
        ["DESPATCH :", metadata.get("DESPATCH", ""), "", "", "COMPLETED CARGO :", metadata.get("Completed Cargo", "")],
        ["LAYTIME TO START COUNTING :", laytime_allowed_value],
        ["TIME ALLOWED :", time_allowed]
    ]

    for row in header_rows:
        ws.append(row)

    ws.append([])  # Spacer row

    # -- Combined Chronological Event and Deduction Table --
    # Adjusted headers to reflect that these are *deducted* events
    ws.append(["Date", "Day", "From", "To", "Deductions (HH:MM)", "To Count (HH:MM)", "Deduction Reason"])
    
    # Apply bold font to header
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    deduction_sum = 0
    to_count_sum = 0
    prev_date = None
    prev_day = None
    for d in deductions:
        date = str(d.get("Date", "")).strip()
        day = str(d.get("Day", "")).strip()

        # Show date/day only if different from previous row
        date_to_write = date if date != prev_date else ""
        day_to_write = day if day != prev_day else ""
        raw_hours = d.get('total_hours')
        try:
            hrs_float = float(raw_hours)
        except (TypeError, ValueError):
            hrs_float = 0.0

        if d.get("deduct"):
            deduction_sum += hrs_float
            ws.append([
                date_to_write,
                day_to_write,
                d.get("deducted_from", ""),
                d.get("deducted_to", ""),
                float_to_hhmm(hrs_float),
                "",
                d.get("Remark", "")
        ])
        else:
            to_count_sum += hrs_float
            ws.append([
                date_to_write,
                day_to_write,
                d.get("deducted_from", ""),
                d.get("deducted_to", ""),
                "",
                float_to_hhmm(hrs_float),
                ""
        ])
        # Update previous date/day
        prev_date = date
        prev_day = day

    ws.append(["TIME ALLOWED :", time_allowed, "", "Total", float_to_hhmm(deduction_sum), float_to_hhmm(to_count_sum)])

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    time_used = f"{net_laytime_used_hours/24.0:.4f}"
    ws.append(["TIME USED", time_used])

    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal="left", vertical="top") # Ensure alignment for this row too
        cell.font = Font(bold=True)

    difference = round(float(time_used) - float(time_allowed), 4)
    if difference > 0:
        rate = float(metadata.get("DEMMURAGE", 0))
        cost = difference * rate
        ws.append(["DEMMURAGE", f"{difference:.4f}"])
        ws.append(["Rate US$", rate, f"{cost:.2f}"])
    elif difference < 0:
        des_pull = abs(difference)
        rate = float(metadata.get("DESPATCH", 0))
        credit = des_pull * rate
        ws.append(["DESPATCH", f"{des_pull:.4f}"])
        ws.append(["Rate US$", rate, f"{credit:.2f}"])
    else:
        ws.append(["NO DEMURRAGE OR DESPATCH APPLICABLE", "0"])

    # Bold all the last few rows
    for row in ws.iter_rows(min_row=ws.max_row - 3, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(bold=True)

    # -- Align all cells left --
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top")

    return wb
